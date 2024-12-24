import io
import json
import base64
from openpyxl import load_workbook
from aws_lambda_powertools.event_handler.api_gateway import ApiGatewayResolver, ProxyEventType, Response

app = ApiGatewayResolver(proxy_type=ProxyEventType.APIGatewayProxyEventV2)

@app.post("/decrypt")
def decrypt_file():
    try:
        # Parse the JSON body from the event
        request = json.loads(app.current_event.body)
    except Exception:
        return Response(
            status_code=400,
            content_type="application/json",
            body=json.dumps({"error": "Invalid JSON format"})
        )

    # Ensure 'file_data' is present in the request
    if 'file_data' not in request:
        return Response(
            status_code=400,
            content_type="application/json",
            body=json.dumps({"error": "Missing 'file_data' base64-encoded in request"})
        )

    file_data = request['file_data']

    try:
        # Decode the base64 file data back into bytes
        excel_content = io.BytesIO(base64.b64decode(file_data))
        
        # Load the workbook
        wb = load_workbook(excel_content)
    except Exception as e:
        return Response(
            status_code=400,
            content_type="application/json",
            body=json.dumps({"error": f"Invalid Excel file: {str(e)}"})
        )

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.password = "your_password"

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    # Encode the resulting workbook in base64 to return as JSON
    result_file_data = base64.b64encode(virtual_workbook.read()).decode()

    return Response(
        status_code=200,
        content_type="application/json",
        body=json.dumps({"file_data": result_file_data})
    )

@app.get("/test")
def test_route():
    return Response(
        status_code=200,
        content_type="application/json",
        body=json.dumps({"message": "Route is working!"})
    )

def lambda_handler(event, context):
    #print
    #return app.resolve(event, context)
    return
    {
        'statusCode': 200,
        'body': json.dumps('hello from github action')
    }
